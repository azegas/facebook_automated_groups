import pyautogui                # automation library
import time                     # time library
import pyperclip                # copy paste allowing library
from openpyxl import Workbook, load_workbook # excel library

excel_file = load_workbook('facebook_groups.xlsx')
excel_sheet = excel_file['pkn']

def main():
    count = 0
    for row in excel_sheet.iter_rows():
        # opens new tab in chrome
        time.sleep(1)
        pyautogui.keyDown('ctrl')
        pyautogui.keyDown('t')
        pyautogui.keyUp('t')
        pyautogui.keyUp('ctrl')
        group_url = row[1].value  # fetch group id from excel
        link = 'https://facebook.com/groups/'+str(group_url) # pro
        # type group name
        pyperclip.copy(link)
        pyautogui.hotkey('ctrl', 'v') # paste
        pyautogui.typewrite('\n')
        count +=1 # variable will increment every loop iteration

    print("Atidariau" + " " + str(count) + " " + "grupes.")
   
if __name__ == '__main__':
	main()        
