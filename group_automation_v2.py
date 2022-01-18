import pyautogui                # automation library
import time                     # time library
import pyperclip                # copy paste allowing library
from openpyxl import Workbook, load_workbook # excel library

# Global variables
# pyautogui.PAUSE = 2.5 # Set up a pause after each PyAutoGUI call

BGBLACK = '\u001b[40m'
BGGREEN = '\u001b[42m'
BGRED = '\u001b[41m'
CEND = '\033[0m'

# posto nuoroda in an an external file
with open('content/nuoroda.txt', 'r') as file:
    nuoroda = file.read().rstrip('\n')

# posto nuoroda in an an external file
with open('content/tekstas.txt', 'r') as file:
    tekstas = file.read().rstrip('\n')

excel_file = load_workbook('facebook_groups.xlsx')
excel_sheet = excel_file['test5']

def main():
   # some time to prepare the browser
   print("3 Seconds to prepare the browser")
   for i in range(4):
      time.sleep(1)
      print("Prepare browser" + " " + str(i) + "/3")
   time.sleep(1)

   print("Opening browser")
   time.sleep(1)
   print("\n")
   open_tab();
   count = 0
   scriptoPradzia = time.time()
   for row in excel_sheet.iter_rows():
      postoPradzia = time.time()
      group_url = row[1].value  # fetch group id from excel
      group_name = row[0].value # fetch group name from excel
      link = 'https://facebook.com/groups/'+str(group_url) # pro
      # type group name
      time.sleep(3)
      pyperclip.copy(link)
      pyautogui.hotkey('ctrl', 'v') # paste
      pyautogui.typewrite('\n')
      print("Atsidariau" + " " + BGBLACK + str(group_name) + CEND)
      print("\n")
      # let browser window load
      print("9 seconds to let the browser window load")
      for i in range(10):
         time.sleep(1)
         print("Browser window load" + " " + str(i) + "/9")
      open_writing_window();
      write_content();
      prepare_next();
      print(BGGREEN + "PAPOSTINTA"+ CEND+ " i " + " " + BGBLACK + str(group_name)+"."+" " + BGRED + "Uztruko {0} sekundes" .format(time.time() - postoPradzia) + CEND)
      print("--------------------------------------------------------------------------------------")
      print("\n")       # new line
      count +=1 # variable will increment every loop iteration
   print("Papostinau i" + " " + str(count) + " " + "grupes.") # how many groups I have posted to 
   print("Is viso uztruko {0} sekundes" .format(time.time() - scriptoPradzia)) # how long it took for the script to run
      
def open_tab():
    # opens new tab in chrome
    pyautogui.keyDown('ctrl')
    pyautogui.keyDown('t')
    pyautogui.keyUp('t')
    pyautogui.keyUp('ctrl')

def open_writing_window():
   try:
      x, y = pyautogui.locateCenterOnScreen("/home/arvydas/Dropbox/projects/facebook_automated_groups/resources/cpp.png", region=(1013, 636, 400, 400))
      print("The image 'create_public_post.png' was found.")
      pyautogui.click(x,y)
   except TypeError:
      print("Could not locate the image - Create a public post...")
      a, b = pyautogui.locateCenterOnScreen("/home/arvydas/Dropbox/projects/facebook_automated_groups/resources/ws.png", region=(1013, 636, 400, 400))
      print("The image 'write something' was found")
      pyautogui.click(a,b)
   time.sleep(2)
      
def write_content():
   pyperclip.copy(nuoroda)
   time.sleep(1)
   pyautogui.hotkey('ctrl', 'v') # paste
   pyautogui.hotkey('ctrl','a') # select all
   pyautogui.press('backspace') # link not necessary anymore, delete
   pyperclip.copy(tekstas)
   pyautogui.hotkey('ctrl', 'v') # paste
   pyautogui.press('enter')      # newline

   time.sleep(1)
   # half(left) screen Acer Aspire V3 771G
   pyautogui.click(1664, 517) # turn off url link according to your screen pixel location (xdotool on linux)
   time.sleep(1)
   pyautogui.click(1448, 950) # Click POST according to your screen pixel location (xdotool on linux)
    
    # # whole screen Lenovo x61s
    # pyautogui.click(751, 478) # turn off url link according to your screen pixel location (xdotool on linux)
    # pyautogui.click(533, 666) # Click POST according to your screen pixel location (xdotool on linux)

def prepare_next():
   # some time to prepare the browser
   print("2 Seconds to prepare the browser")
   for i in range(2):
      time.sleep(1)
      print("Posting" + " " + str(i) + "/2")
      time.sleep(1)
   pyautogui.write(['f6'])     # mark search field, prepare for new link input
      
if __name__ == '__main__':
	main()
